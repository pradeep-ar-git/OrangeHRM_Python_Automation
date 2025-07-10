import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return (sheet.max_row)

def getColumnCount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return (sheet.max_column)

def readData(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num,column_num).value

def writeData(file, sheet_name, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_num,column_num).value = data
    workbook.save(file)

def fillGreenColour(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(row_num, column_num).fill = greenFill
    workbook.save(file)

def fillRedColour(file, sheet_name, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    redFill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(row_num, column_num).fill = redFill
    workbook.save(file)

